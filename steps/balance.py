import openpyxl
from behave import *
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from utilitiy import config_reader

# ---------------------------------------------------------------------------

@given(u'Click on deposit tab')
def step_impl(context):
    # Deposit menu button ---------------------------------------------
    dep=config_reader.Read_config("locators-balance", "depos_PARTIAL_LINK_TEXT")

    context.driver.find_element(By.PARTIAL_LINK_TEXT, 'Deposit').click()
    time.sleep(3)

@given(u'Enter the account number for deposit')
def step_impl(context):
    # account no

    # Read guru excel file:
    workbook = openpyxl.load_workbook("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")
    sheetName = workbook["balance"]

    for i in range(2, sheetName.max_row + 1):
        for j in range(1, 2):
            # for not printing the None
            context.accountNo = sheetName.cell(row=i, column=j).value
            return context.accountNo

    workbook.save("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")


    AccNO = config_reader.Read_config("locators-balance", "AccNO_NAME")

    context.driver.find_element(By.NAME, 'accountno').send_keys(context.contents_1)

@given(u'Enter the amount for deposit')
def step_impl(context):
    # amount
    amm=config_reader.Read_config("locators-balance", "ammount_NAME")
    context.driver.find_element(By.NAME, 'ammount').send_keys(10000)

@when(u'Enter the Description for deposit')
def step_impl(context):
    # description
    desc=config_reader.Read_config("locators-balance", "desc_NAME")
    context.driver.find_element(By.NAME, 'desc').send_keys("deposit 1000 rupee")
    time.sleep(3)

@then(u'Click on submit button for deposit')
def step_impl(context):
    # submit
    sub6=config_reader.Read_config("locators-balance", "sub6_NAME")

    context.driver.find_element(By.NAME, 'AccSubmit').click()
    time.sleep(3)

# ------------------------------------------------------------------------------------

@given(u'Click on withdrawal tab')
def step_impl(context):
    # Withdrawal
    withdraw=config_reader.Read_config("locators-balance", "")

    context.driver.find_element(By.PARTIAL_LINK_TEXT, 'Withdrawal')
    time.sleep(3)

@given(u'Enter the account number for withdrawal')
def step_impl(context):
    # account no
    # Read guru excel file:
    workbook = openpyxl.load_workbook("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")
    sheetName = workbook["balance"]

    for i in range(2, sheetName.max_row + 1):
        for j in range(1, 2):
            # for not printing the None
            context.accountNo = sheetName.cell(row=i, column=j).value
            return context.accountNo

    workbook.save("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")

    Acc_no=config_reader.Read_config("locators-balance", "Acc_no_NAME1")
    context.driver.find_element(By.NAME, 'accountno').send_keys(context.contents_1)
    time.sleep(3)

@given(u'Enter the amount for withdrawal')
def step_impl(context):
    # amount
    withd=config_reader.Read_config("locators-balance", "")
    context.driver.find_element(By.NAME, 'ammount').send_keys(10000)


@when(u'Enter the Description for withdrawal')
def step_impl(context):
    # description
    descr=config_reader.Read_config("locators-balance", "")
    context.driver.find_element(By.NAME, 'desc').send_keys("deposit1")

@then(u'Click on submit button for withdrawal')
def step_impl(context):
    # submit
    sub7=config_reader.Read_config("locators-balance", "")
    context.driver.find_element(By.NAME, 'AccSubmit').click()
    time.sleep(3)

@given(u'Click on Balance Enquiry tab')
def step_impl(context):
    # Balance Enquiry
    bal=config_reader.Read_config("locators-balance", "")
    context.driver.find_element(By.PARTIAL_LINK_TEXT, 'Balance Enquiry').click()
    time.sleep(3)

@when(u'Enter account number for Balance Enquiry')
def step_impl(context):

    # Read guru excel file:
    workbook = openpyxl.load_workbook("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")
    sheetName = workbook["balance"]

    for i in range(2, sheetName.max_row + 1):
        for j in range(1, 2):
            # for not printing the None
            context.accountNo = sheetName.cell(row=i, column=j).value
            return context.accountNo

    workbook.save("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")

    accno=config_reader.Read_config("locators-balance", "Acc_no_NAME3")
    context.driver.find_element(By.NAME, 'accountno').send_keys(context.contents_1)
    time.sleep(3)

@then(u'Click on Submit button for Balance Enquiry')
def step_impl(context):
    sub8=config_reader.Read_config("locators-balance", "sub8_NAME")
    context.driver.find_element(By.NAME, 'AccSubmit').click()
    time.sleep(3)

@given(u'Click on Log out tab')
def step_impl(context):
    # Log out
    logout=config_reader.Read_config("locators-balance", "logout_PARTIAL_LINK_TEXT")
    context.driver.find_element(By.PARTIAL_LINK_TEXT, 'Log out').click()
    time.sleep(3)

@then(u'Switch to alert box')
def step_impl(context):
    # log out alert
    alrt = context.driver.switch_to.alert
    time.sleep(3)
    alrt.accept()
    time.sleep(3)













