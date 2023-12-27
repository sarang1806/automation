from behave import *
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from utilitiy import config_reader,db_function
import openpyxl
import unittest
# from steps import common
import mysql.connector
from mysql.connector import Error

from utilitiy.db_function import get_connection


# ---------------------------------------------------------------------------------------
@given(u'Click on the new customer tab')
def customer_tab(context):
    # New_Customer:

    new_cust= config_reader.Read_config("locators-customer", "new_cust_XPATH")

    context.driver.find_element(By.XPATH, '/html/body/div[3]/div/ul/li[2]/a').click()
    time.sleep(10)


@given(u'Enter the Customer Name "{cname}"')
def customer_name(context, cname):
    context.customer_name,context.date_0f_birth,context.customer_adrdress,context.city,context.state,context.pin,context.mobile_number,context.email_id,context.customer_password= db_function.get_custinfo("2")
    # context.customer_name = db_function.get_custinfo("1")
    cus_name= config_reader.Read_config("locators-customer", "cname_NAME")

    context.driver.find_element(By.NAME, 'name').send_keys(context.customer_name)

@given(u'Select Gender')
def gender(context):
    gender= config_reader.Read_config("locators-customer", "gender_XPATH")

    context.driver.find_element(By.XPATH, '/html/body/table/tbody/tr/td/table/tbody/tr[5]/td[2]/input[2]').click()


@given(u'Select Date of Birth "{dob}"')
def dob(context, dob):
    # context.date_0f_birth = db_function.get_custinfo("1")

    dob= config_reader.Read_config("locators-customer", "dob_NAME")
    context.driver.find_element(By.NAME, 'dob').send_keys(context.date_0f_birth)

@when(u'Enter the Cust Address "{add}"')
def address(context, add):
    # context.customer_adrdress = db_function.get_custinfo("1")

    add= config_reader.Read_config("locators-customer", "add_NAME")

    context.driver.find_element(By.NAME, 'addr').send_keys(context.customer_adrdress)


@when(u'Enter the city "{city}"')
def city(context, city):
    # context.city = db_function.get_custinfo("1")

    city= config_reader.Read_config("locators-customer", "city_NAME")
    context.driver.find_element(By.NAME, 'city').send_keys(context.city)
    time.sleep(2)

@when(u'Enter the state "{state}"')
def state(context, state):
    # context.state = db_function.get_custinfo("1")

    state= config_reader.Read_config("locators-customer", "state_NAME")

    context.driver.find_element(By.NAME, 'state').send_keys(context.state)
    time.sleep(3)

@when(u'Enter the PIN "{pin}"')
def pin(context, pin):
    # context.pin = db_function.get_custinfo("1")

    pin= config_reader.Read_config("locators-customer", "pin_NAME")
    context.driver.find_element(By.NAME, 'pinno').send_keys(context.pin)
    time.sleep(3)

@then(u'Enter the mobile number "{mobile}"')
def mobile_number(context, mobile):
    # context.mobile_number = db_function.get_custinfo("1")

    mob= config_reader.Read_config("locators-customer", "tel_NAME")
    context.driver.find_element(By.NAME, 'telephoneno').send_keys(context.mobile_number)
    time.sleep(3)

@then(u'Enter email-id "{emailid}"')
def email_id(context, emailid):
    # context.email_id = db_function.get_custinfo("1")

    email= config_reader.Read_config("locators-customer", "email_NAME")
    context.driver.find_element(By.NAME, 'emailid').send_keys(context.email_id)
    time.sleep(3)

@then(u'Enter password "{password_2}"')
def password1(context, password_2):
    # context.customer_password = db_function.get_custinfo("1")

    pass2= config_reader.Read_config("locators-customer", "password_NAME")

    context.driver.find_element(By.NAME, 'password').send_keys(context.customer_password)

@then(u'Click on the submit button of new customer section')
def sub_button(context):
    sub1= config_reader.Read_config("locators-customer", "sub_but1_XPATH")
    context.driver.find_element(By.XPATH, '/html/body/table/tbody/tr/td/table/tbody/tr[14]/td[2]/input[1]').click()
    time.sleep(5)

@then(u'Verify customer registration')
def cust_reg(context):
    reg= config_reader.Read_config("locators-customer", "reg_XPATH")

    reg=context.driver.find_element(By.XPATH,'//*[@id="customer"]/tbody/tr[1]/td/p').is_displayed()
    assert reg is True


@then(u'Capture the customer ID and store in the variable')
def custID_variable(context):
    # customer_id :
    cust_id= config_reader.Read_config("locators-customer", "custID_XPATH")

    context.cus_id = context.driver.find_element(By.XPATH, '//*[@id="customer"]/tbody/tr[4]/td[2]').text
    time.sleep(5)
    val = context.cus_id

    # write database
    db_function.custUpdate(customer_id="1",val=val)
    db_function.custUpdate(id="1",val=val)

    # write guru excel file:
    # workbook = openpyxl.load_workbook("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")
    # sheetName=workbook["customer"]

    # c = 1
    # r= 2
    # for r in range (2,r+1):
    #
    #     dict1 = {'customer_id': context.cus_id}
        # to take out the length of the list
        # cnt = len(dict1)

        # to write the data in the cells
        # sheetName.cell(row=r, column=c).value = context.cus_id

        # Logic for arranging the data (list data) in the table in row format
        # for i in range(2, cnt + 2):
        #     sheetName.cell(row=i, column=1).value = dict['customer_id']

    # workbook.save("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")



# --------------------------------------------------------------------------------------------
'''
@given(u'Click on the Edit customer tab of edit customer section')
def step_impl(context):

    # Edit_customer:
    edit_cus= config_reader.Read_config("locators-customer", "editCust_PARTIAL_LINK_TEXT")

    context.driver.find_element(By.PARTIAL_LINK_TEXT,'Edit Customer').click()
    time.sleep(5)

    # Read guru excel file:
    workbook = openpyxl.load_workbook("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")
    sheetName = workbook["customer"]

    for i in range(2, sheetName.max_row + 1):
        for j in range(1,2):
            # for not printing the None
            context.cusid= sheetName.cell(row=i, column=j).value
            return context.cusid

    workbook.save("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")



    cusid= config_reader.Read_config("locators-customer", "cusid_NAME")

    context.driver.find_element(By.NAME, 'cusid').send_keys(context.cusid)
    time.sleep(3)


@then(u'Click on the submit button of edit section')
def step_impl(context):
    sub2= config_reader.Read_config("locators-customer", "sub2_NAME")
    context.driver.find_element(By.NAME, 'AccSubmit').click()
    time.sleep(5)

@when(u'For Editing enter the Address')
def step_impl(context):

    context.driver.find_element(By.NAME, 'addr').clear()
    context.driver.find_element(By.NAME, 'addr').send_keys('Civil lines')
    context.driver.find_element(By.NAME, 'pinno').clear()
    context.driver.find_element(By.NAME, 'telephoneno').clear()
    context.driver.find_element(By.NAME, 'pinno').send_keys('440034')
    context.driver.find_element(By.NAME, 'telephoneno').send_keys('8482653695')

    time.sleep(2)
    context.driver.execute_script("window.scrollBy(0,800)")                        #scroll the window screen


@then(u'Click on the submit button of new edit section')
def step_impl(context):
    sub3= config_reader.Read_config("locators-customer", "sub3_NAME")
    context.driver.find_element(By.NAME,'sub').click()
    time.sleep(3)

@then(u'Switch to alert')
def step_impl(context):
    alrt = context.driver.switch_to.alert
    time.sleep(5)
    alrt.accept()
    time.sleep(5)

#-----------------------------------------------------------------------------------------------
#
# @given(u'Click on the Delete customer tab of delete customer section')
# def step_impl(context):
#
#     #Delete_customer:
#     context.driver.find_element(By.PARTIAL_LINK_TEXT,'Delete Customer').click()
#
# @when(u'Enter the customer ID of delete customer section')
# def step_impl(context):
#     # Read Excel file:
#     f = open("/guru99_cust_id.txt", "r")
#     if f.mode == "r":
#         context.contents = f.read()
#
#     context.driver.find_element(By.NAME,'cusid').send_keys(context.contents)
#
# @then(u'Click on the submit button of delete customer section')
# def step_impl(context):
#     context.driver.find_element(By.NAME,'AccSubmit').click()
#     time.sleep(5)
#
#     alrt = context.driver.switch_to.alert
#     time.sleep(5)
#     alrt.accept()
#     time.sleep(5)

'''










