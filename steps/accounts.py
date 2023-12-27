import openpyxl
from behave import *
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from utilitiy import  config_reader


# ----------------------------------------------------------------------------------------

# new account:
@then(u'click on the new account')
def step_impl(context):
    new_ACC= config_reader.Read_config("locators-account", "newAcc_LINK_TEXT")

    context.driver.find_element(By.LINK_TEXT, 'New Account').click()
    time.sleep(5)

    # Read guru excel file:
    workbook = openpyxl.load_workbook("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")
    sheetName = workbook["customer"]

    for i in range(2, sheetName.max_row + 1):
        for j in range(1, 2):
            # for not printing the None
            context.cusid = sheetName.cell(row=i, column=j).value
            return context.cusid

    workbook.save("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")

    cusid1= config_reader.Read_config("locators-account", "cusid1_NAME")
    context.driver.find_element(By.NAME, 'cusid').send_keys(context.contents_1)
    time.sleep(3)


@then(u'fill complete details')
def step_impl(context):
    sel= config_reader.Read_config("locators-account", "sel_NAME")
    sel = context.driver.find_element(By.NAME, "selaccount")
    drdop = Select(sel)
    drdop.select_by_value("Current")

    deposit= config_reader.Read_config("locators-account", "deposit_NAME")
    context.driver.find_element(By.NAME, "inideposit").send_keys("250000")

@then(u'click on submit button')
def step_impl(context):
    sub4= config_reader.Read_config("locators-account", "sub4_NAME")
    context.driver.find_element(By.NAME, "button2").click()
    time.sleep(5)

    context.accountNo = context.driver.find_element(By.XPATH, '//*[@id="account"]/tbody/tr[4]/td[2]').text

    # write guru excel file:
    workbook = openpyxl.load_workbook("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")
    sheetName = workbook["balance"]

    c = 1
    r = 2
    for r in range(2, r + 1):

        dict2 = {'account_No': context.accountNo}

        # to take out the length of the list
        cnt = len(dict2)

        # to write the data in the cells
        sheetName.cell(row=r, column=c).value = context.cus_id

        # Logic for arranging the data (list data) in the table in row format
        for i in range(2, cnt + 2):
            sheetName.cell(row=i, column=1).value = dict2['account_No']

    workbook.save("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")



## edit account-------------
#
# @then(u'click on the edit account')
# def step_impl(context):
#
#     context.driver.find_element(By.LINK_TEXT, 'Edit Account').click()
#     time.sleep(5)
#
# @then(u'enter the account no.')
# def step_impl(context):

#     # Read Excel file:
#     f = open("D:\\automation_pract\guru99_cust_id.txt", "r")
#     if f.mode == "r":
#         context.contents_1 = f.read()
#
#     context.driver.find_element(By.NAME, "accountno").send_keys(context.contents_1)
#     time.sleep(2)
#
# @then(u'click the submit button')
# def step_impl(context):
#     context.driver.find_element(By.NAME, "AccSubmit").click()
#     time.sleep(3)



# delete account ---------------------------------------------------------------------------------

@then(u'click on the delete account')
def step_impl(context):
    delacc= config_reader.Read_config("locators-account", "delACC_PARTIAL_LINK_TEXT")

    context.driver.find_element(By.PARTIAL_LINK_TEXT, 'Delete Account').click()
    time.sleep(5)

@then(u'enter the account no. for deleted the account')
def step_impl(context):

    # Read guru excel file:
    workbook = openpyxl.load_workbook("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")
    sheetName = workbook["balance"]

    for i in range(2, sheetName.max_row + 1):
        for j in range(1, 2):
            # for not printing the None
            context.accountNo = sheetName.cell(row=i, column=j).value
            return context.accountNo

    workbook.save("C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx")

    accNO= config_reader.Read_config("locators-account", "accNo_NAME")
    context.driver.find_element(By.NAME, "accountno").send_keys(context.contents_1)
    time.sleep(2)



@then(u'click the submit button for delete account')
def step_impl(context):
    sub5= config_reader.Read_config("locators-account", "sub5_NAME")

    context.driver.find_element(By.NAME, 'AccSubmit').click()
    time.sleep(5)
