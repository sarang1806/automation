import openpyxl
path = "D:\\Book2.xlsx"
sheet_name = "Sheet1"
def get_data_from_excel(path,sheet_name):
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]

    total_rows = sheet.max_row
    total_col = sheet.max_column

    # print(sheet.cell(3,1).value)

    for r in range(2,total_rows+1):
        row_list = []
        for c in range(1,total_col+1):
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)
        print(final_list)
    return final_list

