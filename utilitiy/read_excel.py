import openpyxl
workbook = openpyxl.load_workbook("D:\\Book2.xlsx")
sheet_name = workbook["Sheet1"]
path = "D:\\Book2.xlsx"
sheet_name = "Sheet1"
# how many rows in given sheet
def get_row_count(path,sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row

# how many columns in given sheet
def get_column_count(path,sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column


# if you give the row no , column no, form the particular sheet cell data
def get_cell_data(path,sheet_name,row_number,column_number):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.cell(row=row_number,column=column_number).value

def get_data_from_excel(path,sheet_name):

    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    for r in range(2,total_rows+1):
        row_list = []
        for c in range(1,total_cols+1):
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)

    return final_list
