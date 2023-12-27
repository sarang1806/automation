import openpyxl
from utilitiy.config_reader import Read_config

file_path = Read_config("basic-info","XL_path")

workbook = openpyxl.load_workbook(file_path)
# sheetName=workbook["common"]

# file="D:\\seleniumautomation\\BDDhybrid\\guru99_data_xl.xlsx"
# file_path = Read_config("basic-info","XL_path")
#

# To get max number of row:------------------------------------------------------------------------------------------
def getRowCount(file):
    # workbook=openpyxl.load_workbook(file)
    # sheet=workbook.get_sheet_by_name(sheetName)
    sheet=workbook["common"]

    print(sheet.max_row)
    return sheet.max_row

row_count=getRowCount (file="D:\\seleniumautomation\\BDDhybrid\\guru99_data_xl.xlsx")



# To get max number of column:------------------------------------------------------------------------------------------
def getColumnCount(file):
      # workbook=openpyxl.load_workbook(file)
      # sheet=workbook.get_sheet_by_name(sheetName)
      sheet = workbook["common"]

      print(sheet.max_column)
      return sheet.max_column

get_col=getColumnCount(file="D:\\seleniumautomation\\BDDhybrid\\guru99_data_xl.xlsx")




# To get name of the column:-----------------------------------------------------------------------------------------
def getColumn(file):
    # workbook=openpyxl.load_workbook(file, col_name)
    # sheet=workbook.get_sheet_by_name(sheetName)
    sheet = workbook["common"]
    col = sheet['A1']

    print(col.value)
    return col.value

col = getColumn (file="D:\\seleniumautomation\\BDDhybrid\\guru99_data_xl.xlsx")




# To get name of the row:-----------------------------------------------------------------------------------------
def getRow(file):
    # workbook=openpyxl.load_workbook(file)
    # sheet=workbook.get_sheet_by_name(sheetName)
    sheet = workbook["common"]
    row=sheet['A1']
    print(row.value)

    return (row.value)

row=getRow(file="D:\\seleniumautomation\\BDDhybrid\\guru99_data_xl.xlsx")



# To get matched column:-----------------------------------------------------------------------------------------
def getMatchedColumn(file):
    # workbook=openpyxl.load_workbook(file)
    # sheet=workbook.get_sheet_by_name(sheetName)
    sheet = workbook["common"]
    matched_col =sheet['A1']

    print(matched_col.value)
    return (matched_col.value)
match_col=getMatchedColumn(file="D:\\seleniumautomation\\BDDhybrid\\guru99_data_xl.xlsx")




# To get matched row:-----------------------------------------------------------------------------------------
def getMatchedRow(file):
    # workbook=openpyxl.load_workbook(file)
    # sheet=workbook.get_sheet_by_name(sheetName)
    sheet = workbook["common"]
    matched_row = sheet['A1']

    print(matched_row.value)
    return (matched_row.value)

match_row = getMatchedColumn(file="D:\\seleniumautomation\\BDDhybrid\\guru99_data_xl.xlsx")



# To read data in the cell using row and column:------------------------------------------------------------------------
def readData(file,rowNo,columnNo):
    sheet = workbook["common"]

    print(sheet.cell(row=rowNo, column=columnNo).value)
    return sheet.cell(row=rowNo, column=columnNo).value

read_data=readData(file="D:\\seleniumautomation\\BDDhybrid\\guru99_data_xl.xlsx", rowNo=3, columnNo=1)



# To get required row:------------------------------------------------------------------------
# def getRequiredRow (file, col_add, cellvalue):
#     sheet = workbook["common"]
#     sheet.cellvalue="admin"
#
# req_row = getRequiredRow (file="C:\\Users\\DELL\\PycharmProjects\\BDD_Hybrid\\guru99_data_xl.xlsx", col_add=1, cellvalue="admin")




def writeData(file, rowNo, columnNo):
    sheet = workbook["common"]
    data = sheet['A5']
    data.value= "Tester"

    print(sheet.cell(row=rowNo, column=columnNo).value)
    return sheet.cell(row=rowNo, column=columnNo).value
    workbook.save(file)

write_data= writeData (file="D:\\seleniumautomation\\BDDhybrid\\guru99_data_xl.xlsx", rowNo=5,  columnNo=1)

def get_username_password(sheetname,role):
    sheet = workbook[sheetname]
    un = ""
    passward = ""
    max_row = sheet.max_row
    max_col = sheet.max_column
    for i in range(2,max_row+1):
        if sheet.cell(row=i,column=1).value==role:
            un = sheet.cell(row=i,column=2).value
            passward = sheet.cell(row=i,column=3).value
            return un,passward

    if un == "":
        return un,passward


